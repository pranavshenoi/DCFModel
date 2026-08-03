"""
DCF Valuation Tool
==================
A single-file discounted cash flow (DCF) valuation model.

What it does:
  1. Pulls live financial statements for any public ticker (via yfinance)
  2. Calculates WACC from first principles (CAPM + market data)
  3. Projects free cash flow across Bull / Base / Bear scenarios
  4. Discounts those cash flows back to a present intrinsic value per share
  5. Builds a WACC x terminal-growth sensitivity table
  6. Exports everything into a clean, formatted Excel workbook

How to run:
  1. Install the requirements:  pip install -r requirements.txt
  2. Open this file and edit the settings in the "RUN A VALUATION" block
     at the very bottom of the file.
  3. Run:  python dcf_valuation.py
  4. Find the Excel file it prints out at the end.

Everything lives in this one file on purpose — no folders to navigate,
no imports to break. Just edit the ticker at the bottom and run it.

For educational purposes only. This is not financial advice.
"""

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


# ──────────────────────────────────────────────────────────────────────
# STEP 1 — FETCH FINANCIAL DATA
# ──────────────────────────────────────────────────────────────────────

def fetch_company_data(ticker: str) -> dict:
    """Pull income statement, balance sheet, cash flow, and company info."""
    print(f"Fetching data for {ticker}...")
    stock = yf.Ticker(ticker)

    income_stmt = stock.income_stmt
    balance_sheet = stock.balance_sheet
    cash_flow = stock.cashflow
    info = stock.info

    if income_stmt.empty or balance_sheet.empty or cash_flow.empty:
        raise ValueError(
            f"Could not fetch complete financial data for '{ticker}'. "
            f"Check the ticker symbol is correct and try again."
        )

    return {
        "ticker": ticker.upper(),
        "income_stmt": income_stmt,
        "balance_sheet": balance_sheet,
        "cash_flow": cash_flow,
        "info": info,
    }


def calculate_historical_fcf(data: dict) -> pd.Series:
    """Free Cash Flow = Operating Cash Flow - CapEx, for each available year."""
    cf = data["cash_flow"]

    op_cf_row = next((r for r in cf.index if "Operating Cash Flow" in r), None)
    capex_row = next((r for r in cf.index if "Capital Expenditure" in r), None)

    if op_cf_row is None or capex_row is None:
        raise ValueError("Could not locate Operating Cash Flow / CapEx rows.")

    fcf = cf.loc[op_cf_row] + cf.loc[capex_row]  # CapEx is stored as negative
    return fcf.sort_index()


def calculate_tax_rate(data: dict) -> float:
    """Effective tax rate = Tax Provision / Pre-tax Income, most recent year."""
    inc = data["income_stmt"]
    tax_row = next((r for r in inc.index if "Tax Provision" in r), None)
    pretax_row = next((r for r in inc.index if "Pretax Income" in r), None)

    if tax_row is None or pretax_row is None:
        return 0.21  # sensible fallback: US statutory corporate rate

    tax = inc.loc[tax_row].iloc[0]
    pretax = inc.loc[pretax_row].iloc[0]

    if pretax == 0 or pd.isna(pretax) or pd.isna(tax):
        return 0.21

    rate = tax / pretax
    return min(max(rate, 0.0), 0.40)  # clamp to a sane 0-40% range


# ──────────────────────────────────────────────────────────────────────
# STEP 2 — WACC (Weighted Average Cost of Capital)
# ──────────────────────────────────────────────────────────────────────

def safe_num(value, default: float) -> float:
    """Return `value` as a float, or `default` if it's missing/None/NaN.
    Needed because `value or default` does NOT catch NaN (NaN is truthy
    in Python), and min()/max() silently propagate NaN instead of
    clamping it."""
    if value is None:
        return default
    try:
        value = float(value)
    except (TypeError, ValueError):
        return default
    return default if np.isnan(value) else value


def clamp(value: float, low: float, high: float, default: float) -> float:
    """Clip `value` into [low, high], falling back to `default` if it's
    missing or NaN."""
    value = safe_num(value, default)
    return min(max(value, low), high)


def calculate_wacc(
    ticker: str,
    data: dict,
    tax_rate: float,
    risk_free_rate: float = None,
    equity_risk_premium: float = None,
) -> dict:
    """
    WACC = (E/V * Cost of Equity) + (D/V * Cost of Debt * (1 - tax rate))
    Cost of Equity via CAPM: Rf + Beta * Equity Risk Premium

    risk_free_rate / equity_risk_premium: pass these in explicitly for
    non-US markets (e.g. India) — the auto-fetched ^TNX is a US Treasury
    yield and is NOT appropriate as a risk-free rate for other countries.
    If left as None, the function auto-detects US vs India from the
    ticker suffix (.NS / .BO) and picks a sensible regional default.
    """
    info = data["info"]
    bs = data["balance_sheet"]

    is_indian_ticker = ticker.upper().endswith((".NS", ".BO"))

    if risk_free_rate is None:
        if is_indian_ticker:
            # India 10Y G-Sec yield isn't reliably available via yfinance.
            # Uses a reasonable recent-range default — override this
            # yourself with the current rate from rbi.org.in for accuracy.
            risk_free_rate = 0.069  # ~India 10Y G-Sec, approx as of 2026
        else:
            try:
                tnx = yf.Ticker("^TNX").history(period="5d")["Close"].iloc[-1]
                risk_free_rate = tnx / 100
            except Exception:
                risk_free_rate = 0.045  # fallback ~4.5%

    if equity_risk_premium is None:
        # India's equity risk premium runs higher than the US given
        # emerging-market risk; ~6.5-7% is a commonly used range.
        equity_risk_premium = 0.07 if is_indian_ticker else 0.055

    beta = safe_num(info.get("beta"), default=1.0)

    cost_of_equity = risk_free_rate + beta * equity_risk_premium

    # Market cap = equity value
    market_cap = safe_num(info.get("marketCap"), default=0.0)
    if market_cap <= 0:
        shares = safe_num(info.get("sharesOutstanding"), default=0.0)
        price = safe_num(info.get("currentPrice"), default=0.0)
        market_cap = shares * price

    # Total debt from balance sheet
    debt_row = next((r for r in bs.index if "Total Debt" in r), None)
    total_debt = safe_num(bs.loc[debt_row].iloc[0] if debt_row is not None else 0, default=0.0)

    # Approximate cost of debt from interest expense / total debt
    inc = data["income_stmt"]
    interest_row = next((r for r in inc.index if "Interest Expense" in r), None)
    interest_expense = safe_num(
        abs(inc.loc[interest_row].iloc[0]) if interest_row is not None else None,
        default=0.0,
    )
    if total_debt > 0 and interest_expense > 0:
        cost_of_debt = clamp(interest_expense / total_debt, 0.02, 0.12, default=risk_free_rate + 0.02)
    else:
        cost_of_debt = risk_free_rate + 0.02  # fallback spread over risk-free

    total_value = market_cap + total_debt
    weight_equity = market_cap / total_value if total_value > 0 else 1.0
    weight_debt = total_debt / total_value if total_value > 0 else 0.0

    wacc = (weight_equity * cost_of_equity) + (
        weight_debt * cost_of_debt * (1 - tax_rate)
    )
    wacc = clamp(wacc, 0.04, 0.20, default=0.09)  # sanity clamp: 4%-20%, fallback 9%

    return {
        "wacc": wacc,
        "risk_free_rate": risk_free_rate,
        "equity_risk_premium": equity_risk_premium,
        "beta": beta,
        "cost_of_equity": cost_of_equity,
        "cost_of_debt": cost_of_debt,
        "market_cap": market_cap,
        "total_debt": total_debt,
        "weight_equity": weight_equity,
        "weight_debt": weight_debt,
        "tax_rate": tax_rate,
    }


# ──────────────────────────────────────────────────────────────────────
# STEP 3 — DCF PROJECTION
# ──────────────────────────────────────────────────────────────────────

def project_fcf(base_fcf: float, growth_rate: float, years: int = 5) -> list:
    """Project FCF forward, growth tapering slightly each year toward maturity."""
    projections = []
    fcf = base_fcf
    for year in range(1, years + 1):
        # Taper growth 10% of the way toward zero each year (simple fade)
        year_growth = growth_rate * (1 - 0.1 * (year - 1))
        fcf = fcf * (1 + year_growth)
        projections.append(fcf)
    return projections


def discount_cash_flows(cash_flows: list, discount_rate: float) -> list:
    return [cf / ((1 + discount_rate) ** (i + 1)) for i, cf in enumerate(cash_flows)]


def terminal_value(final_year_fcf: float, wacc: float, terminal_growth: float) -> float:
    """Gordon Growth terminal value at the end of the projection period."""
    return (final_year_fcf * (1 + terminal_growth)) / (wacc - terminal_growth)


def run_dcf_scenario(
    base_fcf: float,
    growth_rate: float,
    wacc: float,
    terminal_growth: float,
    net_debt: float,
    shares_outstanding: float,
    years: int = 5,
) -> dict:
    fcf_projections = project_fcf(base_fcf, growth_rate, years)
    discounted_fcf = discount_cash_flows(fcf_projections, wacc)

    tv = terminal_value(fcf_projections[-1], wacc, terminal_growth)
    discounted_tv = tv / ((1 + wacc) ** years)

    enterprise_value = sum(discounted_fcf) + discounted_tv
    equity_value = enterprise_value - net_debt
    intrinsic_value_per_share = equity_value / shares_outstanding if shares_outstanding else 0

    return {
        "fcf_projections": fcf_projections,
        "discounted_fcf": discounted_fcf,
        "terminal_value": tv,
        "discounted_terminal_value": discounted_tv,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "intrinsic_value_per_share": intrinsic_value_per_share,
    }


def run_all_scenarios(data: dict, wacc_result: dict, base_growth: float) -> dict:
    """Run Bull / Base / Bear cases around the base growth assumption."""
    info = data["info"]
    bs = data["balance_sheet"]

    hist_fcf = calculate_historical_fcf(data)
    base_fcf = hist_fcf.iloc[-1]

    debt_row = next((r for r in bs.index if "Total Debt" in r), None)
    cash_row = next((r for r in bs.index if r == "Cash And Cash Equivalents"), None)
    total_debt = bs.loc[debt_row].iloc[0] if debt_row is not None else 0
    cash = bs.loc[cash_row].iloc[0] if cash_row is not None else 0
    net_debt = (total_debt or 0) - (cash or 0)

    shares_outstanding = safe_num(info.get("sharesOutstanding"), default=0.0)
    terminal_growth = 0.025  # long-run GDP-ish growth assumption

    scenarios = {
        "Bull Case": base_growth + 0.05,
        "Base Case": base_growth,
        "Bear Case": max(base_growth - 0.06, 0.0),
    }

    results = {}
    for name, growth in scenarios.items():
        results[name] = run_dcf_scenario(
            base_fcf=base_fcf,
            growth_rate=growth,
            wacc=wacc_result["wacc"],
            terminal_growth=terminal_growth,
            net_debt=net_debt,
            shares_outstanding=shares_outstanding,
        )
        results[name]["growth_rate"] = growth

    return {
        "scenarios": results,
        "base_fcf": base_fcf,
        "net_debt": net_debt,
        "shares_outstanding": shares_outstanding,
        "terminal_growth": terminal_growth,
        "current_price": safe_num(info.get("currentPrice"), default=0.0),
    }


# ──────────────────────────────────────────────────────────────────────
# STEP 4 — SENSITIVITY ANALYSIS
# ──────────────────────────────────────────────────────────────────────

def sensitivity_table(
    base_fcf: float,
    growth_rate: float,
    net_debt: float,
    shares_outstanding: float,
    base_wacc: float,
) -> pd.DataFrame:
    """Intrinsic value per share across a grid of WACC x terminal growth."""
    wacc_range = [base_wacc + d for d in (-0.02, -0.01, 0.0, 0.01, 0.02)]
    growth_range = [0.015, 0.02, 0.025, 0.03, 0.035]

    table = pd.DataFrame(index=[f"{w:.1%}" for w in wacc_range],
                          columns=[f"{g:.1%}" for g in growth_range])

    for w in wacc_range:
        for g in growth_range:
            if w <= g:
                value = None  # not mathematically valid (WACC must exceed g)
            else:
                result = run_dcf_scenario(
                    base_fcf, growth_rate, w, g, net_debt, shares_outstanding
                )
                value = result["intrinsic_value_per_share"]
            table.loc[f"{w:.1%}", f"{g:.1%}"] = value

    return table


def implied_growth_rate(
    market_price: float,
    base_fcf: float,
    net_debt: float,
    shares_outstanding: float,
    wacc: float,
    terminal_growth: float,
) -> float:
    """Binary-search the FCF growth rate that makes intrinsic value = market price."""
    low, high = -0.10, 0.40
    for _ in range(60):
        mid = (low + high) / 2
        result = run_dcf_scenario(
            base_fcf, mid, wacc, terminal_growth, net_debt, shares_outstanding
        )
        if result["intrinsic_value_per_share"] < market_price:
            low = mid
        else:
            high = mid
    return (low + high) / 2


# ──────────────────────────────────────────────────────────────────────
# STEP 5 — EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────────

NAVY = "0A1628"
GOLD = "C9A227"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"


def _style_header(cell, fill=NAVY, font_color=WHITE, bold=True, size=11):
    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.font = Font(color=font_color, bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def export_to_excel(data: dict, wacc_result: dict, scenario_data: dict,
                     sensitivity: pd.DataFrame, implied_growth: float) -> str:
    wb = Workbook()
    ticker = data["ticker"]

    # ---------- Sheet 1: Cover ----------
    ws = wb.active
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    ws["A1"] = f"DCF Valuation — {ticker}"
    ws["A1"].font = Font(size=18, bold=True, color=NAVY, name="Calibri")
    ws.merge_cells("A1:C1")
    ws["A2"] = f"Generated {datetime.now().strftime('%d %b %Y')}"
    ws["A2"].font = Font(size=10, italic=True, color="777777")
    ws.merge_cells("A2:C2")

    ws["A4"] = "Scenario"
    ws["B4"] = "FCF Growth"
    ws["C4"] = "Intrinsic Value/Share"
    for col in ("A4", "B4", "C4"):
        _style_header(ws[col])

    row = 5
    for name, result in scenario_data["scenarios"].items():
        ws[f"A{row}"] = name
        ws[f"B{row}"] = result["growth_rate"]
        ws[f"B{row}"].number_format = "0.00%"
        ws[f"C{row}"] = result["intrinsic_value_per_share"]
        ws[f"C{row}"].number_format = '"$"#,##0.00'
        for col in ("A", "B", "C"):
            ws[f"{col}{row}"].border = _thin_border()
        row += 1

    row += 1
    ws[f"A{row}"] = "Current Market Price"
    ws[f"B{row}"] = scenario_data["current_price"]
    ws[f"B{row}"].number_format = '"$"#,##0.00'
    ws[f"A{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "Implied FCF Growth (Market)"
    ws[f"B{row}"] = implied_growth
    ws[f"B{row}"].number_format = "0.00%"
    ws[f"A{row}"].font = Font(bold=True)

    row += 2
    ws[f"A{row}"] = "WACC Breakdown"
    ws[f"A{row}"].font = Font(bold=True, color=GOLD, size=12)
    row += 1
    for label, key, fmt in [
        ("WACC", "wacc", "0.00%"),
        ("Cost of Equity", "cost_of_equity", "0.00%"),
        ("Cost of Debt", "cost_of_debt", "0.00%"),
        ("Beta", "beta", "0.00"),
        ("Risk-Free Rate", "risk_free_rate", "0.00%"),
        ("Tax Rate", "tax_rate", "0.00%"),
    ]:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = wacc_result[key]
        ws[f"B{row}"].number_format = fmt
        row += 1

    # ---------- Sheet 2: DCF Model ----------
    ws2 = wb.create_sheet("DCF Model")
    ws2.column_dimensions["A"].width = 24
    for col in "BCDEFG":
        ws2.column_dimensions[col].width = 15

    ws2["A1"] = f"{ticker} — Free Cash Flow Projections"
    ws2["A1"].font = Font(size=14, bold=True, color=NAVY)
    ws2.merge_cells("A1:F1")

    base_row = 3
    for name, result in scenario_data["scenarios"].items():
        ws2[f"A{base_row}"] = name
        ws2[f"A{base_row}"].font = Font(bold=True, color=GOLD, size=12)
        base_row += 1

        ws2[f"A{base_row}"] = "Year"
        for i in range(1, 6):
            cell = ws2.cell(row=base_row, column=1 + i, value=f"Year {i}")
            _style_header(cell, fill=NAVY, size=10)
        _style_header(ws2[f"A{base_row}"], fill=NAVY, size=10)
        base_row += 1

        ws2[f"A{base_row}"] = "Projected FCF"
        for i, val in enumerate(result["fcf_projections"]):
            c = ws2.cell(row=base_row, column=2 + i, value=val)
            c.number_format = '"$"#,##0,,"M"'
        base_row += 1

        ws2[f"A{base_row}"] = "Discounted FCF"
        for i, val in enumerate(result["discounted_fcf"]):
            c = ws2.cell(row=base_row, column=2 + i, value=val)
            c.number_format = '"$"#,##0,,"M"'
        base_row += 1

        ws2[f"A{base_row}"] = "Terminal Value (PV)"
        ws2[f"B{base_row}"] = result["discounted_terminal_value"]
        ws2[f"B{base_row}"].number_format = '"$"#,##0,,"M"'
        base_row += 1

        ws2[f"A{base_row}"] = "Enterprise Value"
        ws2[f"B{base_row}"] = result["enterprise_value"]
        ws2[f"B{base_row}"].number_format = '"$"#,##0,,"M"'
        base_row += 1

        ws2[f"A{base_row}"] = "Equity Value"
        ws2[f"B{base_row}"] = result["equity_value"]
        ws2[f"B{base_row}"].number_format = '"$"#,##0,,"M"'
        base_row += 1

        ws2[f"A{base_row}"] = "Intrinsic Value / Share"
        ws2[f"B{base_row}"] = result["intrinsic_value_per_share"]
        ws2[f"B{base_row}"].number_format = '"$"#,##0.00'
        ws2[f"A{base_row}"].font = Font(bold=True)
        ws2[f"B{base_row}"].font = Font(bold=True, color=GOLD)
        base_row += 2

    # ---------- Sheet 3: Sensitivity ----------
    ws3 = wb.create_sheet("Sensitivity")
    ws3["A1"] = f"{ticker} — Sensitivity: Intrinsic Value/Share"
    ws3["A1"].font = Font(size=14, bold=True, color=NAVY)
    ws3.merge_cells("A1:G1")
    ws3["A2"] = "Rows: WACC   |   Columns: Terminal Growth Rate"
    ws3["A2"].font = Font(italic=True, size=9, color="777777")

    start_row = 4
    ws3.cell(row=start_row, column=1, value="WACC \\ g")
    _style_header(ws3.cell(row=start_row, column=1))
    for j, col_name in enumerate(sensitivity.columns):
        c = ws3.cell(row=start_row, column=2 + j, value=col_name)
        _style_header(c)

    for i, idx in enumerate(sensitivity.index):
        r = start_row + 1 + i
        c = ws3.cell(row=r, column=1, value=idx)
        _style_header(c, fill=LIGHT_GREY, font_color=NAVY)
        for j, col_name in enumerate(sensitivity.columns):
            val = sensitivity.loc[idx, col_name]
            cell = ws3.cell(row=r, column=2 + j)
            if val is None:
                cell.value = "n/a"
            else:
                cell.value = val
                cell.number_format = '"$"#,##0.00'
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(sensitivity.columns) + 2):
        ws3.column_dimensions[get_column_letter(col)].width = 14

    # ---------- Save ----------
    output_dir = os.path.join(os.path.expanduser("~"), "Documents", "Project", "Python Projects")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"{ticker}_valuation.xlsx")
    wb.save(filepath)
    return filepath


# ──────────────────────────────────────────────────────────────────────
# PUTTING IT ALL TOGETHER
# ──────────────────────────────────────────────────────────────────────

def run_valuation(
    ticker: str,
    manual_growth: float = None,
    risk_free_rate: float = None,
    equity_risk_premium: float = None,
) -> str:
    """
    End-to-end DCF valuation for any public ticker.

    manual_growth: overrides the historical FCF growth rate used as the
        Base Case assumption (e.g. 0.10 for 10%). Leave as None to
        estimate it from historical free cash flow trends instead.

    risk_free_rate / equity_risk_premium: override the CAPM inputs used
        for WACC. For Indian tickers (.NS / .BO) these auto-default to
        India-appropriate values, but you should pass the current India
        10Y G-Sec yield yourself for accuracy — check rbi.org.in or a
        financial data provider for the latest figure.
        Example for an Indian stock:
            run_valuation("TCS.NS", manual_growth=0.12, risk_free_rate=0.069)
    """
    print("=" * 60)
    print("  DCF VALUATION TOOL")
    print(f"  Ticker: {ticker.upper()}")
    print("=" * 60)

    data = fetch_company_data(ticker)
    currency = data["info"].get("currency", "USD")
    tax_rate = calculate_tax_rate(data)
    wacc_result = calculate_wacc(
        ticker, data, tax_rate,
        risk_free_rate=risk_free_rate,
        equity_risk_premium=equity_risk_premium,
    )

    if manual_growth is not None:
        base_growth = manual_growth
    else:
        hist_fcf = calculate_historical_fcf(data)
        if len(hist_fcf) >= 2 and hist_fcf.iloc[0] > 0:
            years_span = len(hist_fcf) - 1
            cagr = (hist_fcf.iloc[-1] / hist_fcf.iloc[0]) ** (1 / years_span) - 1
            base_growth = min(max(cagr, 0.0), 0.25)  # clamp to a sane range
        else:
            base_growth = 0.08  # fallback assumption

    scenario_data = run_all_scenarios(data, wacc_result, base_growth)

    sensitivity = sensitivity_table(
        base_fcf=scenario_data["base_fcf"],
        growth_rate=base_growth,
        net_debt=scenario_data["net_debt"],
        shares_outstanding=scenario_data["shares_outstanding"],
        base_wacc=wacc_result["wacc"],
    )

    implied_g = implied_growth_rate(
        market_price=scenario_data["current_price"],
        base_fcf=scenario_data["base_fcf"],
        net_debt=scenario_data["net_debt"],
        shares_outstanding=scenario_data["shares_outstanding"],
        wacc=wacc_result["wacc"],
        terminal_growth=scenario_data["terminal_growth"],
    )

    filepath = export_to_excel(data, wacc_result, scenario_data, sensitivity, implied_g)

    print("\nResults:")
    for name, result in scenario_data["scenarios"].items():
        print(f"  {name:<12} growth={result['growth_rate']:.1%}  "
              f"value=${result['intrinsic_value_per_share']:.2f}")
    print(f"\n  Current market price: ${scenario_data['current_price']:.2f}")
    print(f"  Implied market growth: {implied_g:.1%}")
    print("\n" + "=" * 60)
    print("  VALUATION COMPLETE")
    print(f"  Output saved to: {filepath}")
    print("=" * 60)

    return filepath


# ──────────────────────────────────────────────────────────────────────
# RUN A VALUATION — edit the line below and run this file
# ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    run_valuation("AAPL", manual_growth=0.10)  # <-- change ticker/growth here
